#!/usr/bin/env python3
"""
phase20-electrode-coverage-audit.py — SAFETY-CRITICAL READ-ONLY audit of the
JM Die electrode + taptite corpus against the "Automated Program" tracking
workbook (TRAINING-LEARNING-MS0 / MS0-U3 companion script).

NOTE: this filename intentionally shares the `phase20-*` prefix with
`phase20-lathe-template-corpus-scan.py` (U1 sibling). The suffixes are
distinct (`-lathe-template-corpus-scan` vs `-electrode-coverage-audit`)
and they live in the same Docustrata index dir — no collision, just
shared scan-era prefix. This is intentional per the spec MS0-U3 unit-id
`U-TL-U3-ELECTRODE-COVERAGE-AUDIT` which names the file literally.

Inputs (all READ-ONLY):
    --index   Docustrata jm-die-index-v2.json (default: HERE/jm-die-index-v2.json)
    --xlsm    H:/PRISM/JM DIE/Automated Program_Corrected 5-25.xlsm
    --corpus  H:/PRISM/JM DIE (fallback when --index unusable)

Outputs (atomic write via tempfile + rename, never overwrites in-place):
    --out     mcp-server/data/training/audits/electrode-coverage/_audit-snapshot.json

Output schema (consumed by ElectrodeCoverageAuditEngine.report({presetSnapshot})):
    {
      "schemaVersion": 1,
      "generatedAt": "<ISO timestamp>",
      "corpusScan": {
        "electrodeCount": int,
        "taptiteCount": int,
        "samplePaths": {"electrodes": [...], "taptites": [...]},
        "byTopDir": {"electrodes": {...}, "taptites": {...}},
        "byExtension": {"electrodes": {...}, "taptites": {...}},
        "source": "phase20-electrode-coverage-audit.py",
      },
      "xlsmFingerprint": {
        "exists": bool,
        "path": str,
        "mtimeMs": float,
        "sizeBytes": int,
        "sha256": str,
      },
      "xlsmContent": {
        "rowCount": int,
        "sheetNames": [str, ...],
        "programIds": [str, ...],   # parsed from workbook column A (if openpyxl available)
        "parsedWith": "openpyxl|skipped"
      }
    }

Safety guarantees (verified by the engine's mtime-invariance test):
    - Opens .xlsm strictly READ-ONLY (openpyxl read_only=True; falls back to
      `skipped` mode if openpyxl is unavailable — the engine's xlsm fingerprint
      is still emitted)
    - Never writes/renames/chmods any corpus file or the .xlsm.
    - Never follows symlinks during the corpus walk.
    - Atomic snapshot write: tempfile in same dir → os.replace into --out path.

Usage:
    python phase20-electrode-coverage-audit.py
    python phase20-electrode-coverage-audit.py --dry-run
    python phase20-electrode-coverage-audit.py --xlsm <path> --out <path>
    python phase20-electrode-coverage-audit.py --max-depth 6
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import tempfile
import time
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterator


HERE = Path(__file__).resolve().parent
PRISM = HERE.parent.parent  # H:/prism

DEFAULT_INDEX = HERE / "jm-die-index-v2.json"
DEFAULT_XLSM = Path("H:/PRISM/JM DIE/Automated Program_Corrected 5-25.xlsm")
DEFAULT_CORPUS = Path("H:/PRISM/JM DIE")
DEFAULT_OUT = (
    PRISM
    / "mcp-server"
    / "data"
    / "training"
    / "audits"
    / "electrode-coverage"
    / "_audit-snapshot.json"
)

SAMPLE_PATHS_PER_CATEGORY = 8
DEFAULT_MAX_DEPTH = 8
SHA256_CHUNK = 64 * 1024
MAX_XLSM_FINGERPRINT_BYTES = 200 * 1024 * 1024
SCHEMA_VERSION = 1

ELECTRODE_RE = re.compile(r"electrode", re.IGNORECASE)
TAPTITE_RE = re.compile(r"taptite", re.IGNORECASE)


# ───────────────────────────────────────────────────────────────────────────────
# Corpus walk — depth-limited, symlink-safe.

def walk_corpus(root: Path, max_depth: int) -> dict[str, Any]:
    """Walk `root` to `max_depth`; collect electrode + taptite files.

    Returns dict with: electrodes[], taptites[], dirsVisited, truncated, errors[].
    Never throws on EACCES/etc — collects in `errors[]` instead.
    """
    electrodes: list[Path] = []
    taptites: list[Path] = []
    errors: list[str] = []
    dirs_visited = 0
    truncated = False

    stack: list[tuple[Path, int]] = [(root, 0)]
    while stack:
        dir_path, depth = stack.pop()
        if depth > max_depth:
            truncated = True
            continue
        try:
            entries = list(os.scandir(dir_path))
        except (OSError, PermissionError) as e:
            errors.append(f"{dir_path}: {e}")
            continue
        dirs_visited += 1
        for ent in entries:
            try:
                if ent.is_symlink():
                    continue  # never follow symlinks
                if ent.is_dir(follow_symlinks=False):
                    stack.append((Path(ent.path), depth + 1))
                elif ent.is_file(follow_symlinks=False):
                    name = ent.name
                    if ELECTRODE_RE.search(name):
                        electrodes.append(Path(ent.path))
                    elif TAPTITE_RE.search(name):
                        taptites.append(Path(ent.path))
            except OSError as e:
                errors.append(f"{ent.path}: {e}")

    electrodes.sort()
    taptites.sort()
    return {
        "electrodes": electrodes,
        "taptites": taptites,
        "dirsVisited": dirs_visited,
        "truncated": truncated,
        "errors": errors,
    }


def top_dir_of(p: Path, root: Path) -> str:
    try:
        rel = p.relative_to(root)
    except ValueError:
        return "(out-of-root)"
    parts = rel.parts
    return parts[0] if parts else "(root)"


def ext_of(p: Path) -> str:
    return p.suffix.lower() if p.suffix else "(noext)"


def tally(items: list[Path], key_fn) -> dict[str, int]:
    c: Counter[str] = Counter()
    for it in items:
        c[key_fn(it)] += 1
    return dict(c)


# ───────────────────────────────────────────────────────────────────────────────
# Xlsm fingerprint — stat + streaming sha256, never opens for write.

def fingerprint_xlsm(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {
            "ok": True,
            "exists": False,
            "path": str(path),
            "fingerprintedAt": datetime.now(timezone.utc).isoformat(),
        }
    try:
        stat = path.stat()
    except OSError as e:
        return {
            "ok": False,
            "exists": False,
            "path": str(path),
            "error": "xlsm_stat_failed",
            "detail": str(e),
        }
    if not path.is_file():
        return {
            "ok": False,
            "path": str(path),
            "error": "xlsm_not_a_file",
            "detail": f"{path} exists but is not a regular file",
        }
    if stat.st_size > MAX_XLSM_FINGERPRINT_BYTES:
        return {
            "ok": False,
            "path": str(path),
            "error": "xlsm_too_large",
            "detail": f"size {stat.st_size} > {MAX_XLSM_FINGERPRINT_BYTES}",
        }
    h = hashlib.sha256()
    try:
        with path.open("rb") as f:
            while True:
                chunk = f.read(SHA256_CHUNK)
                if not chunk:
                    break
                h.update(chunk)
    except OSError as e:
        return {
            "ok": False,
            "path": str(path),
            "error": "xlsm_read_failed",
            "detail": str(e),
        }
    return {
        "ok": True,
        "exists": True,
        "path": str(path),
        "mtimeMs": stat.st_mtime * 1000.0,
        "sizeBytes": stat.st_size,
        "sha256": h.hexdigest(),
        "fingerprintedAt": datetime.now(timezone.utc).isoformat(),
    }


# ───────────────────────────────────────────────────────────────────────────────
# Xlsm content parse — optional, only when openpyxl is installed.

def parse_xlsm_content(path: Path, max_rows: int = 5000) -> dict[str, Any]:
    """Read .xlsm with openpyxl in read_only mode. NEVER opens with write access."""
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError:
        return {
            "rowCount": 0,
            "sheetNames": [],
            "programIds": [],
            "parsedWith": "skipped",
            "reason": "openpyxl_not_installed",
        }
    if not path.exists():
        return {
            "rowCount": 0,
            "sheetNames": [],
            "programIds": [],
            "parsedWith": "skipped",
            "reason": "xlsm_missing",
        }
    try:
        # read_only=True, keep_vba=False — strictly read-only.
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001 — openpyxl raises a wide variety
        return {
            "rowCount": 0,
            "sheetNames": [],
            "programIds": [],
            "parsedWith": "skipped",
            "reason": f"openpyxl_load_failed: {e}",
        }
    try:
        sheet_names = list(wb.sheetnames)
        row_count = 0
        program_ids: list[str] = []
        for sn in sheet_names:
            try:
                ws = wb[sn]
            except KeyError:
                continue
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= max_rows:
                    break
                row_count += 1
                # Heuristic: column-A non-empty cell = program-id-ish.
                if row and row[0] is not None:
                    val = str(row[0]).strip()
                    if val and len(val) < 80:
                        program_ids.append(val)
        return {
            "rowCount": row_count,
            "sheetNames": sheet_names,
            "programIds": program_ids[:200],  # cap surface to keep snapshot compact
            "parsedWith": "openpyxl",
        }
    finally:
        try:
            wb.close()
        except Exception:  # noqa: BLE001
            pass


# ───────────────────────────────────────────────────────────────────────────────
# Atomic write — tempfile in same dir, then os.replace.

def atomic_write_json(out_path: Path, data: Any) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(
        prefix=out_path.name + ".",
        suffix=".tmp",
        dir=str(out_path.parent),
    )
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2, sort_keys=True)
        os.replace(tmp, out_path)
    except Exception:
        try:
            os.unlink(tmp)
        except OSError:
            pass
        raise


# ───────────────────────────────────────────────────────────────────────────────
# Main.

def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--index", type=Path, default=DEFAULT_INDEX)
    ap.add_argument("--xlsm", type=Path, default=DEFAULT_XLSM)
    ap.add_argument("--corpus", type=Path, default=DEFAULT_CORPUS)
    ap.add_argument("--out", type=Path, default=DEFAULT_OUT)
    ap.add_argument("--max-depth", type=int, default=DEFAULT_MAX_DEPTH)
    ap.add_argument("--no-xlsm-content", action="store_true",
                    help="Skip openpyxl content parse (fingerprint only)")
    ap.add_argument("--dry-run", action="store_true",
                    help="Don't write the snapshot file")
    args = ap.parse_args(argv)

    if args.max_depth < 0 or args.max_depth > 20:
        print(f"phase20: --max-depth must be 0..20 (got {args.max_depth})",
              file=sys.stderr)
        return 2

    # Snapshot the .xlsm mtime BEFORE any read — used to verify invariance below.
    xlsm_mtime_before: float | None = None
    try:
        xlsm_mtime_before = args.xlsm.stat().st_mtime
    except OSError:
        pass

    start = time.time()
    walk = walk_corpus(args.corpus, args.max_depth)
    walk_ms = int((time.time() - start) * 1000)

    fp = fingerprint_xlsm(args.xlsm)
    content = (
        {"rowCount": 0, "sheetNames": [], "programIds": [], "parsedWith": "skipped",
         "reason": "disabled_by_flag"}
        if args.no_xlsm_content
        else parse_xlsm_content(args.xlsm)
    )

    # Verify the .xlsm wasn't touched (parity with the engine's mtime assertion).
    xlsm_mtime_after: float | None = None
    try:
        xlsm_mtime_after = args.xlsm.stat().st_mtime
    except OSError:
        pass
    if xlsm_mtime_before is not None and xlsm_mtime_after is not None:
        if xlsm_mtime_before != xlsm_mtime_after:
            print(
                f"phase20: FATAL — .xlsm mtime changed during audit "
                f"({xlsm_mtime_before} → {xlsm_mtime_after}). Aborting write.",
                file=sys.stderr,
            )
            return 3

    snapshot = {
        "schemaVersion": SCHEMA_VERSION,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "generator": "phase20-electrode-coverage-audit.py",
        "milestone": "TRAINING-LEARNING-MS0/U-TL-U3-ELECTRODE-COVERAGE-AUDIT",
        "corpusScan": {
            "corpusRoot": str(args.corpus),
            "maxDepth": args.max_depth,
            "electrodeCount": len(walk["electrodes"]),
            "taptiteCount": len(walk["taptites"]),
            "dirsVisited": walk["dirsVisited"],
            "truncated": walk["truncated"],
            "walkTimeMs": walk_ms,
            "byTopDir": {
                "electrodes": tally(walk["electrodes"], lambda p: top_dir_of(p, args.corpus)),
                "taptites": tally(walk["taptites"], lambda p: top_dir_of(p, args.corpus)),
            },
            "byExtension": {
                "electrodes": tally(walk["electrodes"], ext_of),
                "taptites": tally(walk["taptites"], ext_of),
            },
            "samplePaths": {
                "electrodes": [str(p) for p in walk["electrodes"][:SAMPLE_PATHS_PER_CATEGORY]],
                "taptites": [str(p) for p in walk["taptites"][:SAMPLE_PATHS_PER_CATEGORY]],
            },
            "errors": walk["errors"][:20],
        },
        "xlsmFingerprint": fp,
        "xlsmContent": content,
    }

    if args.dry_run:
        print(json.dumps(snapshot, indent=2, sort_keys=True))
        return 0

    atomic_write_json(args.out, snapshot)
    print(
        f"phase20: snapshot written to {args.out} — "
        f"electrodes={snapshot['corpusScan']['electrodeCount']}, "
        f"taptites={snapshot['corpusScan']['taptiteCount']}, "
        f"xlsm_exists={fp.get('exists', False)}, "
        f"walk_ms={walk_ms}",
        file=sys.stderr,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
